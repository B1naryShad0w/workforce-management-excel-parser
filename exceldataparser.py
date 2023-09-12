import openpyxl
import math
import pandas as pd
import tkinter as tk
from tkinter import filedialog
from tkinter import simpledialog
from tkinter import messagebox

notes = ""
multiplier = 2.0
hourlyRate = 55.0
df = None
sheet_name = None

def find_column(df, keyword):
    for idx, col in enumerate(df.columns):
        col_name = str(col)
        if keyword.lower() in col_name.lower() or df[col].astype(str).str.contains(keyword, case=False, na=False).any():
            return idx
    return None

def process_data():
    global multiplier, notes
    results_text = ""
    global_multiplier = multiplier  # Store the global multiplier in a separate variable
    global_notes = notes
    for i in range(len(df.iloc[:, 0])):
        if ((type(df.iloc[i, cHr]) != str and df.iloc[i, cHr] > 0 and df.iloc[i, pNUM][:5] != "Total") or (type(df.iloc[i, cEA]) != str and df.iloc[i, cEA] > 0 and df.iloc[i, pNUM][:5] != "Total")):
            current_multiplier = df.iloc[i, mULT] if mULT is not None and df.iloc[i, mULT] > 0 else global_multiplier
            current_notes = str(df.iloc[i, nOTES]) if nOTES is not None and not pd.isna(df.iloc[i, nOTES]) else global_notes
            current_invoice = df.iloc[i, iNV] if iNV is not None and df.iloc[i, iNV] > 0 else 0
            current_newENB = df.iloc[i, pENB] + (df.iloc[i, cLA] * current_multiplier) + df.iloc[i, cEA]
            results_text += df.iloc[i, pNUM] + "\n"
            results_text += "Prior ENB: $" + format(df.iloc[i, pENB], ",.2f") + "\n"
            results_text += "Current Labor Amount: $" + format(df.iloc[i, cLA], ",.2f") + "\n"
            results_text += "Multiplier: " + str(current_multiplier) + "\n"
            results_text += "ENB for this period: $" + format(df.iloc[i, cLA] * current_multiplier, ",.2f") + "\n"
            results_text += "Current Expenses Amount: $" + format(df.iloc[i, cEA], ",.2f") + "\n"
            results_text += "New ENB: $" + format(current_newENB, ",.2f") + "\n"
            results_text += "Invoice Amount for This Period: $" + format(df.iloc[i, iNV], ",.2f") + "\n"
            results_text += "Ending ENB: $" + format(current_newENB - current_invoice, ",.2f") + "\n"
            if current_notes:
                results_text += "Notes: " + current_notes + "\n"
            results_text += "\n"
    results_textbox.delete(1.0, tk.END)
    results_textbox.insert(tk.END, results_text)

def save_to_file():
    global df, multiplier
    global_multiplier = multiplier
    global_rate = hourlyRate

    if df is None:
        messagebox.showerror("No data", "Please load and process data first.")
        return

    new_df = pd.DataFrame(columns=['Project Name', 'Project Manager', 'Client Name', 'Contract Total Compensation', 'Prior ENB', 'New ENB', 'Multiplier', 'Average Hourly Rate', 'Remaining Hours to Complete', 'Invoice Amount for This Period', 'Ending ENB'])

    for i in range(len(df.iloc[:, 0])):
        if (type(df.iloc[i, cTC]) != str and df.iloc[i, cTC] > 0 and df.iloc[i, pNUM][:5] != "Total"):
            current_multiplier = df.iloc[i, mULT] if mULT is not None and df.iloc[i, mULT] > 0 else global_multiplier
            current_rate = df.iloc[i, aHR] if aHR is not None and df.iloc[i, aHR] > 0 else global_rate
            current_invoice = df.iloc[i, iNV] if iNV is not None and df.iloc[i, iNV] > 0 else 0
            new_enb = round(df.iloc[i, pENB] + (df.iloc[i, cLA] * current_multiplier) + df.iloc[i, cEA], 2)
            ending_enb = round(new_enb - current_invoice, 2)

            new_row = {
                'Project Name': df.iloc[i, pNUM],
                'Project Manager': df.iloc[i, pMG],
                'Client Name': df.iloc[i, cNM],
                'Contract Total Compensation': df.iloc[i, cTC],
                'Prior ENB': df.iloc[i, pENB],
                'New ENB': new_enb,
                'Multiplier': current_multiplier,
                'Average Hourly Rate': current_rate,
                'Remaining Hours to Complete': round(((df.iloc[i, cTC] * 0.8) - new_enb) / (current_rate * current_multiplier), 2),
                'Invoice Amount': current_invoice,
                'Ending ENB': ending_enb
            }

            new_df = pd.concat([new_df, pd.DataFrame([new_row])], ignore_index=True)

    save_file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx;*.xls")], confirmoverwrite=False)

    if save_file_path:
        new_sheet_name = simpledialog.askstring("New Sheet", "Enter a new sheet name:", initialvalue=sheet_name)

        if new_sheet_name:
            try:
                with pd.ExcelWriter(save_file_path, mode='a', engine='openpyxl') as writer:
                    if new_sheet_name in writer.book.sheetnames:
                        messagebox.showerror("Sheet name exists", "The sheet name already exists in the selected file. Please try again.")
                    else:
                        new_df.to_excel(writer, sheet_name=new_sheet_name, index=False)
                        messagebox.showinfo("File saved", f"Data saved in sheet '{new_sheet_name}' of file '{save_file_path}'.")
            except FileNotFoundError:
                with pd.ExcelWriter(save_file_path, mode='w', engine='openpyxl') as writer:
                    if new_sheet_name in writer.book.sheetnames:
                        messagebox.showerror("Sheet name exists", "The sheet name already exists in the selected file. Please try again.")
                    else:
                        new_df.to_excel(writer, sheet_name=new_sheet_name, index=False)
                        messagebox.showinfo("File saved", f"Data saved in sheet '{new_sheet_name}' of file '{save_file_path}'.")
            except PermissionError:
                messagebox.showerror("Permission Error", "The file is likely open in another program. Please close the file and try again.")

def open_file():
    global df, file_path, sheet_name
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
    if file_path:
        select_button.config(state=tk.NORMAL)
        select_sheet()

def select_sheet():
    global df, file_path, sheet_name
    # Get the list of sheet names from the chosen file
    workbook = openpyxl.load_workbook(file_path)
    sheet_names = workbook.sheetnames

    # Create a custom dialog box for selecting a sheet name
    def on_sheet_select(event):
        global sheet_name
        sheet_name = sheet_listbox.get(sheet_listbox.curselection())
        sheet_dialog.destroy()

    sheet_dialog = tk.Toplevel(root)
    sheet_dialog.title("Select Sheet")

    sheet_listbox = tk.Listbox(sheet_dialog)
    for name in sheet_names:
        sheet_listbox.insert(tk.END, name)
    sheet_listbox.pack()
    sheet_listbox.bind('<<ListboxSelect>>', on_sheet_select)

    # Wait for the user to select a sheet name
    root.wait_window(sheet_dialog)

    # Read the selected sheet into a DataFrame
    if sheet_name:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        file_label.config(text=f"File: {file_path}\nSheet: {sheet_name}")
        override_columns_button.config(state=tk.NORMAL)
        confirm_and_override_columns()

def check_columns_defined():
    return all(col is not None for col in [pNUM, cHr, pENB, cLA, cEA, cTC, pMG, cNM, jTD])

def get_multiplier_and_notes():
    global multiplier, hourlyRate, notes
    multiplier = simpledialog.askfloat("Multiplier", "Enter multiplier value:", initialvalue=multiplier)
    if multiplier is not None:
        hourlyRate = simpledialog.askfloat("Hourly Rate", "Enter average hourly rate:", initialvalue=hourlyRate)
        if hourlyRate is not None:
            notes = simpledialog.askstring("Notes", "Enter additional notes (optional):", initialvalue=notes)
            multiplier_notes_label.config(text=f"Multiplier: {multiplier}\nHourly Rate: {hourlyRate}\nNotes: {notes}")
            if df is not None and check_columns_defined():
                process_button.config(state=tk.NORMAL)
                new_button.config(state=tk.NORMAL)

def show_preview():
    if df is not None:
        # Create a new window to display the DataFrame
        table_window = tk.Toplevel(root)
        table_window.title("Data Preview")

        # Format the DataFrame as a string with a custom line width
        pd.set_option("display.max_columns", None)
        pd.set_option("display.width", None)
        df_string = df.head(20).to_string()

        # Display the DataFrame as a string in a Text widget
        df_preview = tk.Text(table_window, wrap=tk.NONE, width=120, height=20)
        df_preview.insert(tk.END, df_string)
        df_preview.config(state=tk.DISABLED)
        df_preview.pack(expand=True, fill=tk.BOTH)

        # Add scrollbars
        scrollbar_y = tk.Scrollbar(table_window, orient=tk.VERTICAL, command=df_preview.yview)
        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)
        scrollbar_x = tk.Scrollbar(table_window, orient=tk.HORIZONTAL, command=df_preview.xview)
        scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)
        df_preview.config(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        return table_window
    else:
        messagebox.showerror("No file loaded", "Please load a file first.")

def confirm_and_override_columns():
    global pNUM, cHr, pENB, cLA, cEA, mULT, aHR, cTC, pMG, cNM, jTD, iNV, nOTES
    table_window = None

    pNUM = find_column(df, 'Project Earnings')
    cHr = find_column(df, 'Current\nHours')
    pENB = find_column(df, 'Prior\nENB')
    cLA = find_column(df, 'Cur Labor\nAmount')
    cEA = find_column(df, 'Cur Exp\nAmount')
    mULT = find_column(df, 'Multiplier')
    aHR = find_column(df, 'Average Hourly Rate')
    cTC = find_column(df, 'Contract Total\nCompensation')
    pMG = find_column(df, 'Project\nManager')
    cNM = find_column(df, 'Client\nName')
    jTD = find_column(df, 'JTD\nBilled')
    iNV = find_column(df, 'Invoice Amount')
    nOTES = find_column(df, 'Note')

    if not check_columns_defined():
        override_columns()
        return

    columns_text = f"Project Names: {pNUM}\nCurrent Hours: {cHr}\nPrior ENB: {pENB}\nCur Labor Amount: {cLA}\nCur Exp Amount: {cEA}\nMultiplier: {mULT if mULT is not None else 'Not Found'}\nAvg Hourly Rate: {aHR if aHR is not None else 'Not Found'}\nNotes: {nOTES if nOTES is not None else 'Not Found'}\nTotal Compensation: {cTC}\nProject Manager: {pMG}\nClient Name: {cNM}\nJTD Billed: {jTD}"

    def on_preview_click():
        nonlocal table_window 
        table_window = show_preview()

    def on_yes_click():
        custom_dialog.destroy()
        if df is not None and multiplier is not None and hourlyRate is not None and check_columns_defined():
           process_button.config(state=tk.NORMAL)
           new_button.config(state=tk.NORMAL)

    def on_no_click():
        custom_dialog.response = False
        custom_dialog.destroy()
        override_columns(table_window)

    custom_dialog = tk.Toplevel(root)
    custom_dialog.title("Confirm columns")
    custom_dialog.response = True

    message_label = tk.Label(custom_dialog, text=f"Found columns:\n{columns_text}\n\nIs this correct?")
    message_label.pack()

    preview_button = tk.Button(custom_dialog, text="Show Preview", command=on_preview_click)
    preview_button.pack()

    yes_button = tk.Button(custom_dialog, text="Yes", command=on_yes_click)
    yes_button.pack(side=tk.LEFT, padx=(20, 10))

    no_button = tk.Button(custom_dialog, text="No", command=on_no_click)
    no_button.pack(side=tk.RIGHT, padx=(10, 20))

    root.wait_window(custom_dialog)

def unload_file():
    global df, file_path, sheet_name, pNUM, cHr, pENB, cLA, cEA, cTC, pMG, cNM, jTD, mULT, aHR, nOTES
    df = None
    file_path = None
    sheet_name = None
    pNUM, cHr, pENB, cLA, cEA, cTC, pMG, cNM, jTD, mULT, aHR, nOTES = None, None, None, None, None, None, None, None, None, None, None, None
    file_label.config(text="")
    results_textbox.delete(1.0, tk.END)
    select_button.config(state=tk.DISABLED)
    process_button.config(state=tk.DISABLED)
    new_button.config(state=tk.DISABLED)
    override_columns_button.config(state=tk.DISABLED)

def override_columns(table_window=None):
    global pNUM, cHr, pENB, cLA, cEA, mULT, aHR, cTC, pMG, cNM, jTD, nOTES

    if not table_window:
        table_window = show_preview()

    pNUM = get_valid_column_index("Enter new column index for Project Names:", len(df.columns), pNUM)
    if pNUM is None:
        if table_window:
            table_window.destroy()
        unload_file()
        return
    cHr = get_valid_column_index("Enter new column index for Current Hours:", len(df.columns), cHr)
    if cHr is None:
        if table_window:
            table_window.destroy()
        unload_file()
        return
    pENB = get_valid_column_index("Enter new column index for Prior ENB:", len(df.columns), pENB)
    if pENB is None:
        if table_window:
            table_window.destroy()
        unload_file()
        return
    cLA = get_valid_column_index("Enter new column index for Current Labor Amount:", len(df.columns), cLA)
    if cLA is None:
        if table_window:
            table_window.destroy()
        unload_file()
        return
    cEA = get_valid_column_index("Enter new column index for Current Exp Amount:", len(df.columns), cEA)
    if cEA is None:
        if table_window:
            table_window.destroy()
        unload_file()
        return
    mULT = get_valid_column_index("Enter new column index for Multiplier (cancel for None):", len(df.columns), mULT)
    aHR = get_valid_column_index("Enter new column index for Avg Hourly Rate (cancel for None):", len(df.columns), aHR)
    nOTES = get_valid_column_index("Enter new column index for Notes (cancel for None):", len(df.columns), nOTES)
    cTC = get_valid_column_index("Enter new column index for Contract Total Compensation:", len(df.columns), cTC)
    if cTC is None:
        if table_window:
            table_window.destroy()
        unload_file()
        return
    pMG = get_valid_column_index("Enter new column index for Project Manager:", len(df.columns), pMG)
    if pMG is None:
        if table_window:
            table_window.destroy()
        unload_file()
        return
    cNM = get_valid_column_index("Enter new column index for Client Name:", len(df.columns), cNM)
    if cNM is None:
        if table_window:
            table_window.destroy()
        unload_file()
        return
    jTD = get_valid_column_index("Enter new column index for JTD Billed:", len(df.columns), jTD)
    if jTD is None:
        if table_window:
            table_window.destroy()
        unload_file()
        return
    if df is not None and multiplier is not None and hourlyRate is not None:
        process_button.config(state=tk.NORMAL)
        new_button.config(state=tk.NORMAL)
    
def get_valid_column_index(prompt, max_col, default=None):
    while True:
        col_index = simpledialog.askinteger("Override Column", prompt, parent=root, initialvalue=default)
        if col_index is None or (0 <= col_index < max_col):
            return col_index
        messagebox.showwarning("Invalid Column Index", "Please enter a valid column index.")

def show_help():
    help_text = """Excel Data Parser by Shahnawaz Haque V1.0

This program helps you process data from Excel files.


Import Data\nOpen an Excel file to start.\n
Choose Sheet\nSelect a different sheet within the opened Excel file.\n
Override Columns\nCustomize columns for processing data.\n
Configure Settings\nSet default multiplier, hourly rate, and notes.\n
Calculate Results\nProcess the data and display results.\n
Export Results\nProcess the data and save results to an Excel file.\n
Clear Data\nClear the loaded data and start over.


This is specialized software with a specific use case. Processing of input is predefined."""

    messagebox.showinfo("Help", help_text)

root = tk.Tk()
root.title("Excel Data Parser by Shahnawaz Haque V.1.0")

open_button = tk.Button(root, text="Import Data", command=open_file)
open_button.grid(row=0, column=0, padx=5, pady=5)

select_button = tk.Button(root, text="Choose Sheet", command=select_sheet, state=tk.DISABLED)
select_button.grid(row=0, column=1, padx=5, pady=5)

file_label = tk.Label(root, text="")
file_label.grid(row=1, column=0, padx=5, pady=5, columnspan=2)

override_columns_button = tk.Button(root, text="Configure Columns", command=confirm_and_override_columns, state=tk.DISABLED)
override_columns_button.grid(row=2, column=0, padx=5, pady=5)

get_multiplier_notes_button = tk.Button(root, text="Configure Settings", command=get_multiplier_and_notes)
get_multiplier_notes_button.grid(row=2, column=1, padx=5, pady=5)

multiplier_notes_label = tk.Label(root, text="")
multiplier_notes_label.grid(row=3, column=0, padx=5, pady=5, columnspan=2)
multiplier_notes_label.config(text=f"Multiplier: {multiplier}\nHourly Rate: {hourlyRate}\nNotes: {notes}")

process_button = tk.Button(root, text="Calculate Results", command=process_data, state=tk.DISABLED)
process_button.grid(row=4, column=0, padx=5, pady=5)

new_button = tk.Button(root, text="Export Results", command=save_to_file, state=tk.DISABLED)
new_button.grid(row=4, column=1, padx=5, pady=5)

clear_data_button = tk.Button(root, text="Clear Data", command=unload_file)
clear_data_button.grid(row=5, column=0, padx=5, pady=5)

results_textbox = tk.Text(root, wrap=tk.WORD)
results_textbox.grid(row=0, column=2, rowspan=7, padx=5, pady=5, sticky='nsew')

help_button = tk.Button(root, text="Help", command=show_help)
help_button.grid(row=5, column=1, padx=5, pady=5)

root.grid_columnconfigure(2, weight=1)
root.grid_rowconfigure(6, weight=1)

root.mainloop()
